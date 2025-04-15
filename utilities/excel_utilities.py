import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Any, Optional


class ExcelUtilities:

    @staticmethod
    def load_workbook(file_path: str) -> Optional[Workbook]:
        try:
            return openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            print(f"[!] File not found: {file_path}")
        except Exception as e:
            print(f"[!] Failed to load workbook '{file_path}': {e}")
        return None

    @staticmethod
    def get_sheet(workbook: Workbook, sheet_name: str) -> Optional[Worksheet]:
        try:
            return workbook[sheet_name]
        except KeyError:
            print(f"[!] Sheet '{sheet_name}' not found in workbook.")
        except Exception as e:
            print(f"[!] Failed to access sheet '{sheet_name}': {e}")
        return None

    @staticmethod
    def read_cell(sheet: Worksheet, row: int, column: int) -> Any:
        try:
            return sheet.cell(row=row, column=column).value
        except Exception as e:
            print(f"[!] Failed to read cell at (row={row}, column={column}): {e}")
            return None

    @staticmethod
    def write_cell(sheet: Worksheet, row: int, column: int, value: Any) -> None:
        try:
            sheet.cell(row=row, column=column).value = value
        except Exception as e:
            print(f"[!] Failed to write to cell at (row={row}, column={column}): {e}")

    @staticmethod
    def get_row_count(sheet: Worksheet) -> int:
        try:
            return sheet.max_row
        except Exception as e:
            print(f"[!] Failed to get row count: {e}")
            return 0

    @staticmethod
    def get_column_count(sheet: Worksheet) -> int:
        try:
            return sheet.max_column
        except Exception as e:
            print(f"[!] Failed to get column count: {e}")
            return 0

    @staticmethod
    def read_sheet_as_matrix(sheet: Worksheet) -> List[List[Any]]:
        try:
            return [
                [cell for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        except Exception as e:
            print(f"[!] Failed to read sheet as matrix: {e}")
            return []

    @staticmethod
    def save_workbook(workbook: Workbook, file_path: Optional[str] = None) -> None:
        try:
            workbook.save(file_path)
        except Exception as e:
            print(f"[!] Failed to save workbook: {e}")

    @staticmethod
    def get_column_values(sheet: Worksheet, column: int) -> List[Any]:
        try:
            return [sheet.cell(row=row, column=column).value for row in range(2, sheet.max_row + 1)]
        except Exception as e:
            print(f"[!] Failed to get column values from column {column}: {e}")
            return []

    @staticmethod
    def get_row_values(sheet: Worksheet, row: int) -> List[Any]:
        try:
            return [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        except Exception as e:
            print(f"[!] Failed to get values from row {row}: {e}")
            return []

    @staticmethod
    def find_row_by_cell_value(sheet: Worksheet, column: int, value: Any) -> int:
        try:
            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if cell_value == value:
                    return row
        except Exception as e:
            print(f"[!] Failed to find row with value '{value}' in column {column}: {e}")
        return -1

    @staticmethod
    def copy_sheet(workbook: Workbook, source_sheet: str, new_sheet_name: str) -> None:
        try:
            source = workbook[source_sheet]
            workbook.copy_worksheet(source).title = new_sheet_name
        except Exception as e:
            print(f"[!] Failed to copy sheet '{source_sheet}' to '{new_sheet_name}': {e}")

    @staticmethod
    def delete_sheet(workbook: Workbook, sheet_name: str) -> None:
        try:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
        except Exception as e:
            print(f"[!] Failed to delete sheet '{sheet_name}': {e}")

    @staticmethod
    def create_workbook_with_sheet(file_path: str, sheet_name: str) -> None:
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            workbook.save(file_path)
            print(f"[+] New workbook created: {file_path}")
        except Exception as e:
            print(f"[!] Failed to create new workbook '{file_path}': {e}")
